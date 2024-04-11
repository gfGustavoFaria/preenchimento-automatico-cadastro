import openpyxl # Leitura da planilha excel
import pyperclip # copiar e colar mantendo os acentos nas strings
import pyautogui # movimentar o mouse para copiar e colar
from time import sleep #pausa

# Entrar na planilha
dados_produtos = openpyxl.load_workbook('produtos_ficticios.xlsx') # load_workbook carrega a planilha
pagina_produtos = dados_produtos['Produtos']

# copiar informação de um campo e colar no campo correspondente

for linha in pagina_produtos.iter_rows(min_row=2): #Iterar a partir da linha dois
    nome_produto = linha[0].value
    pyperclip.copy(nome_produto)
    pyautogui.click(338,285,duration=1) # vai até a posição do campo (coordenadas) a ser preenchido na web
    pyautogui.hotkey('ctrl','v') # cola o dado no campo informado acima

    descricao = linha[1].value
    pyperclip.copy(descricao)
    pyautogui.click(176,395,duration=1)
    pyautogui.hotkey('ctrl','v')

    categoria = linha[2].value
    pyperclip.copy(categoria)
    pyautogui.click(178,554,duration=1)
    pyautogui.hotkey('ctrl','v')

    codigo_produto = linha[3].value
    pyperclip.copy(codigo_produto)
    pyautogui.click(174,662,duration=1)
    pyautogui.hotkey('ctrl','v')

    peso_produto = linha[4].value
    pyperclip.copy(peso_produto)
    pyautogui.click(176,769,duration=1)
    pyautogui.hotkey('ctrl','v')

    dimensoes = linha[5].value
    pyperclip.copy(dimensoes)
    pyautogui.click(185,872,duration=1)
    pyautogui.hotkey('ctrl','v')

    pyautogui.click(191,951,duration=1) #Clica no botão "Próximo"
    sleep(3) # Espera 3 segundos para continuar as próximas ações

    preco = linha[6].value
    pyperclip.copy(preco)
    pyautogui.click(183,312,duration=1)
    pyautogui.hotkey('ctrl','v')

    qtd_estoque = linha[7].value
    pyperclip.copy(qtd_estoque)
    pyautogui.click(193,415, duration=1)
    pyautogui.hotkey('ctrl','v')

    data_validade = linha[8].value
    pyperclip.copy(data_validade)
    pyautogui.click(201,532,duration=1)
    pyautogui.hotkey('ctrl','v')

    cor = linha[9].value
    pyperclip.copy(cor)
    pyautogui.click(196,635,duration=1)
    pyautogui.hotkey('ctrl','v')


    tamanho = linha[10].value
    pyautogui.click(180,738,duration=1)
    if tamanho == 'Pequeno':
        pyautogui.click(228,780,duration=1)
    elif tamanho == 'Médio':
        pyautogui.click(240,813,duration=1)
    else:
        pyautogui.click(214,832,duration=1)

    material = linha[11].value
    pyperclip.copy(material)
    pyautogui.click(175,845,duration=1)
    pyautogui.hotkey('ctrl','v')

    pyautogui.click(186,928,duration=1)

    fabricante = linha[12].value
    pyperclip.copy(fabricante)
    pyautogui.click(209,340,duration=1)
    pyautogui.hotkey('ctrl','v')

    pais_origem = linha[13].value
    pyperclip.copy(pais_origem)
    pyautogui.click(182,441,duration=1)
    pyautogui.hotkey('ctrl','v')

    observacoes = linha[14].value
    pyperclip.copy(observacoes)
    pyautogui.click(182,552,duration=1)
    pyautogui.hotkey('ctrl','v')

    codigo_de_barras = linha[15].value
    pyperclip.copy(codigo_de_barras)
    pyautogui.click(181,719,duration=1)
    pyautogui.hotkey('ctrl','v')

    localizacao_no_armazem = linha[16].value
    pyperclip.copy(localizacao_no_armazem)
    pyautogui.click(172,819,duration=1)
    pyautogui.hotkey('ctrl','v')

# Botão "Concluir"
    pyautogui.click(180,899,duration=1)
    # Botão "Confirmar inclusão"  
    pyautogui.click(1169,239,duration=2)
# Botão "adiconar mais 1"    
    pyautogui.click(950,625, duration=2)
    


# Repetir esses passos para outros campos até preencher todos os campos da página
#
# Clicar em próximo
#
# Repetir os mesmos passos e ir para a próxima página
#
# Rpetir os mesmos passos e finalizar o cadastro e clicar em concluir
#
# Clicar em ok, para finalizar o processo
#
# Clicar no ok novamente, na mensagem de confirmação de salvamento no banco de dados
#
# Clicar em "adicionar mais um" e repetir o processo até finalizar a planilha
